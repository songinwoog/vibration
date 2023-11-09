#진동측정값 읽어 오기
import os
import sys

#모드버스 호출 임포트
from pymodbus.client import ModbusTcpClient
from pymodbus.transaction import *
from pymodbus.payload import Endian
from pymodbus.payload import BinaryPayloadBuilder
from pymodbus.payload import BinaryPayloadDecoder
from datetime import datetime
import time

#엑셀 저장을 위한 모듈 임포트
from openpyxl import load_workbook
import pandas as pd

def excel_save(line):
        file_name = 'vibration.xlsx'

        wb = load_workbook(file_name)

        data = wb.active

        last_num = data['A1'].value

        new_num = last_num+1
        data['A1'] = new_num
# print (last_num)
# print(data['A'+str(last_num)].value)
# print(data['A3'].value)
#wb = openpyxl.Workbook()

# ws = wb.create_sheet('진동센서 동작')
        data['A'+str(last_num)] = line_list[line]
        data['B'+str(last_num)] = now.strftime('%Y-%m-%d %H:%M:%S')

        wb.save('vibration.xlsx')






#호기 배열 생성
line_list = ['V151 진공펌프_진동HI발생',
             'V151 2층 RC FAN 진동HI발생',
             'V151 3층 RC FAN 진동HI발생',
             'V151 4층 RC FAN 진동HI발생',
             'V151 5층 RC FAN 진동HI발생',
             'V151 6층 RC FAN 진동HI발생',   
             'V151 7층 RC FAN 진동HI발생',
             'V151 6층돌출 RC FAN 진동HI발생',
             'V152 진공펌프_진동HI발생',   
             'V151 2층 RC FAN 진동HI발생',
             'V152 3층 RC FAN 진동HI발생', 
             'V152 4층 RC FAN 진동HI발생',
             'V201 진공펌프_진동HI발생',   
             'V201 2층 RC FAN 진동HI발생',
             'V201 3층 RC FAN 진동HI발생',
             'V201 4층 RC FAN 진동HI발생',
             'V202 진공펌프_진동HI발생',  
             'V202 2층 RC FAN 진동HI발생',  
             'V202 3층 RC FAN 진동HI발생',
             'V202 4층 RC FAN 진동HI발생',
                                       
             

             ]



#os.system("cls")

#client = ModbusTcpClient('172.16.3.158',502)
client = ModbusTcpClient('127.0.0.1',502)
client.connect


#   time.sleep(0.3)
print('진동센서 관리서버 연결중...')
#배열 초기화
try:
        arrayInit = client.read_holding_registers(0,25)
        first_result = arrayInit
        joinStatus = True
except Exception as e:
       joinStatus = False
       print ('초기화 에러 -> 통신오류')
       #print (e)


while joinStatus:    
        clientResult = client.read_holding_registers(0,25)
        alarm_list_result = clientResult.registers

#print(arr_result)      
        alarmQuest = alarm_list_result.index(1)!=24
        if(alarmQuest):
                alarmNewAction = first_result != alarm_list_result
                if(alarmNewAction):
                    first_result = alarm_list_result
                    save_line = alarm_list_result.index(1)    
                    #print(arr_result.index(1))
                    now = datetime.now()
                    print(alarm_list_result , now.strftime('%Y-%m-%d %H:%M:%S'))
                    excel_save(save_line)

        else:
               first_result = alarm_list_result

        # for n in range(0,24):
        #     if(arr_result[n]==1):
        #         print ('True 발생 : '+str(n))
        #         sys.exit()
time.sleep(0.2)
        #print('True 발견 : ${n}')
# for i in range(0,17):
#     print(result.bits)
    
#result1 = client.read_holding_registers(0,4)

#print(result1(0))
#decode = BinaryPayloadDecoder.fromRegisters(result1.registers,Endian.BIG)
#value = str(round(decode.decode_32bit_float(),1))
#print (value)
#print(result.registers)