
#젊음이 그립당.... 
import openpyxl
from openpyxl import Workbook

#wb = openpyxl.load_workbook('test.xlsx')
write_wb = Workbook()

write_ws = write_wb.create_sheet('생성시트')

write_wb.save('sheet_create')

#wb.save('test2.xlsx')
#wb.close()